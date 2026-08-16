"""Parsing and validation for staged CSV/XLSX schedule imports."""
import csv
import io
import logging
import re
from datetime import date, datetime, time, timedelta

from openpyxl import load_workbook

from app.models import (Division, Field, FieldConfigurationOption, FieldInstance,
                        Game, GameSlot, HostLocation, HostingAvailability,
                        HostLocationConfiguration, PhysicalFieldArea, Season,
                        TimeslotFieldConfiguration, Week)
from app.teams import resolve_roster_team, season_roster
from app.facility_layouts import (JOHNSBURG_APPROVED_LAYOUT_CODES_BY_LOCATION,
                                  johnsburg_field_templates,
                                  johnsburg_location_name)
from app.services.field_resolution import (resolve_active_field,
                                            resolve_legacy_import_field)
from app.services.facility_layout_validation import (get_active_supported_layouts,
                                                      layout_label)

REQUIRED = ('week', 'date', 'kickoff', 'site', 'field', 'fieldtype', 'division', 'hometeam', 'awayteam')
logger = logging.getLogger(__name__)


def _key(value):
    return re.sub(r'[^a-z0-9]', '', str(value or '').strip().lower())


def _text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip())


def _normalized_name(value):
    return _text(value).casefold()


def _normalized_field_type(value):
    """Return a canonical individual field type, independent of layout labels."""
    normalized = _text(value).casefold()
    return normalized.upper() if normalized in {'small', 'medium', 'large'} else None


def _division_alias(value):
    """Normalize known legacy spelling of the configured Coed 6-7 division."""
    text = _text(value).casefold()
    text = re.sub(r'\bco\s*-?\s*ed\b', 'coed', text)
    text = re.sub(r'\b6\s*(?:-|,|/|and)\s*7(?:\s*,\s*8)?\b', '6-7', text)
    return _key(text)


def _split_area_slot(raw):
    area = _text(raw.get('physicalarea'))
    slot = _text(raw.get('field'))
    if not area and '/' in slot:
        area, slot = (_text(part) for part in slot.split('/', 1))
    return area, slot


def _slot_signature(slot_name, field_type):
    match = re.fullmatch(r'(small|medium|large)(?:\s+field)?\s*(\d+)',
                         _text(slot_name), re.IGNORECASE)
    if not match:
        return None
    size = match.group(1).upper()
    return (size, int(match.group(2))) if not field_type or size == field_type else None


def configuration_supports_required_slots(existing_slot_ids, required_slot_ids):
    """Return whether every imported slot is provided by a saved layout.

    Callers deliberately pass canonical ``Field`` IDs (legacy layouts) or
    generated ``GameSlot`` IDs (physical-area layouts), never display names.
    Extra IDs in the saved layout are unused capacity and are therefore valid.
    """
    return set(required_slot_ids) <= set(existing_slot_ids)


def _layout_integrity_error(configuration, site, physical_area=None):
    """Validate the complete persisted membership of a custom field layout.

    Capacity columns are not a substitute for live member records.  In
    particular, silently dropping a deleted member can make a stale layout
    appear to support an import by name while its foreign keys cannot be used.
    """
    if not configuration or not configuration.is_active:
        return 'Configuration could not be resolved or is inactive.'
    members = list(configuration.members or [])
    member_ids = [member.field_id for member in members]
    if len(member_ids) != len(set(member_ids)):
        return 'Configuration contains a duplicate physical field assignment.'
    expected_by_type = {
        size: int(getattr(configuration, f'{size.lower()}_field_count', 0) or 0)
        for size in ('SMALL', 'MEDIUM', 'LARGE')
    }
    if not sum(expected_by_type.values()):
        # Older custom layouts predate the capacity columns. Their canonical
        # label still describes the required complete membership.
        label = _text(configuration.configuration_name).upper().replace('_', ' ')
        for count, size in re.findall(r'(\d+)\s+(SMALL|MEDIUM|LARGE)', label):
            expected_by_type[size] += int(count)
    expected = sum(expected_by_type.values())
    if len(members) != expected:
        return (f'Configuration is incomplete: it requires {expected} physical fields '
                f'but has {len(members)} assigned.')
    for member in members:
        field = member.field
        if not field:
            return ('Configuration contains a reference to a physical field that no longer '
                    'exists. Edit the configuration and reassign its fields before importing.')
        if (field.deleted_at is not None or not field.is_active
                or field.host_location_id != site.id
                or (physical_area is not None
                    and field.physical_field_area_id != physical_area.id)):
            return (f'Configuration references field "{field.name}", but that field is inactive, '
                    'deleted, or no longer assigned to this physical area.')
        field_type = _normalized_field_type(field.layout_type)
        configured_count = expected_by_type.get(field_type, 0)
        actual_count = sum(1 for item in members if item.field and
                           _normalized_field_type(item.field.layout_type) == field_type)
        if not field_type or actual_count > configured_count:
            return f'Configuration references field "{field.name}" with an incompatible field type.'
    return None


def _area_slot_candidates(db, area, slot_name, field_type):
    signature = _slot_signature(slot_name, field_type)
    if not signature:
        return []
    size, index = signature
    return [option for option in db.query(FieldConfigurationOption).filter_by(
        physical_field_area_id=area.id, is_active=True).all()
            if index <= int(getattr(option, f'{size.lower()}_field_count', 0) or 0)]


def _find_area_instance(db, site, area, option, slot_name, game_date, kickoff):
    """Resolve a generated slot only through its physical-area ancestry."""
    query = (db.query(FieldInstance).join(
        HostingAvailability, FieldInstance.hosting_availability_id == HostingAvailability.id
    ).filter(FieldInstance.host_location_id == site.id,
             FieldInstance.instance_date == game_date,
             FieldInstance.is_active.is_(True),
             HostingAvailability.physical_field_area_id == area.id))
    instances = query.all()
    requested = _normalized_name(slot_name)
    matches = [item for item in instances
               if (_normalized_name(item.field_name) == requested
                   or _normalized_name(item.field_name).endswith(' ' + requested))]
    if option:
        option_matches = [item for item in matches if
                          item.hosting_availability.field_configuration_option_id == option.id]
        if option_matches:
            matches = option_matches
    timed = [item for item in matches if db.query(GameSlot).filter_by(
        field_instance_id=item.id, slot_date=game_date, start_time=kickoff).first()]
    matches = timed or matches
    return matches[0] if len(matches) == 1 else None


def _area_hosting_availability(db, season_id, site, area, game_date, kickoff):
    """Return area-level availability, independently of generated resources."""
    return (db.query(HostingAvailability).filter(
        HostingAvailability.season_id == season_id,
        HostingAvailability.host_location_id == site.id,
        HostingAvailability.physical_field_area_id == area.id,
        HostingAvailability.available_date == game_date,
        HostingAvailability.active.is_(True),
        HostingAvailability.is_available.is_(True),
        HostingAvailability.start_time <= kickoff,
        HostingAvailability.end_time > kickoff,
    ).order_by(HostingAvailability.start_time.desc()).first())


def _week_number(value):
    """Parse spreadsheet numbers and human-readable labels such as ``Week 1``."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float) and value.is_integer():
        return int(value) if value > 0 else None
    match = re.fullmatch(r'(?:week\s*)?(\d+)', _text(value), re.IGNORECASE)
    number = int(match.group(1)) if match else None
    return number if number and number > 0 else None


def parse_schedule_file(filename: str, content: bytes):
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if suffix == 'csv':
        rows = list(csv.reader(io.StringIO(content.decode('utf-8-sig'))))
    elif suffix == 'xlsx':
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = [list(row) for row in book.active.iter_rows(values_only=True)]
    else:
        raise ValueError('Only .csv and .xlsx schedule files are supported.')
    if not rows:
        raise ValueError('The import file is empty.')
    headers = [_key(value) for value in rows[0]]
    missing = [name for name in REQUIRED if name not in headers]
    if missing:
        raise ValueError('Missing required columns: ' + ', '.join(missing))
    return [{headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            for row in rows[1:] if any(value not in (None, '') for value in row)]


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(_text(value), fmt).date()
        except ValueError:
            pass
    return None


def _time(value):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (float, int)) and 0 <= value < 1:
        return (datetime.min + timedelta(days=float(value))).time().replace(second=0, microsecond=0)
    for fmt in ('%I:%M %p', '%I:%M%p', '%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(_text(value).upper(), fmt).time()
        except ValueError:
            pass
    return None


def _configuration_candidates(db, site, field_name, field_type):
    """Return supported layouts containing this logical position and size.

    A supported facility layout is scheduling metadata in its own right.  It
    must remain usable by imports even when the corresponding layout is not the
    site's active/default ``HostLocationConfiguration`` (the import may be the
    reason the facility is reconfigured for that wave).
    """
    active = get_active_supported_layouts(db, site.id)
    configured_by_code = {
        configuration.configuration_name.strip().upper().replace('-', '_').replace(' ', '_'): configuration
        for configuration in active
    }
    location = johnsburg_location_name(site)
    layout_codes = (JOHNSBURG_APPROVED_LAYOUT_CODES_BY_LOCATION.get(location, ())
                    if location else configured_by_code)
    candidates = {}
    requested_name = _normalized_name(field_name)
    if _normalized_name(getattr(site, 'name', '')) == 'hiller stadium' and requested_name == 'field 2':
        requested_name = 'field 3'
    for layout_code in layout_codes:
        configuration = configured_by_code.get(layout_code)
        templates = johnsburg_field_templates(site, layout_code)
        if configuration and configuration.members:
            templates = [(member.field.name, member.field.layout_type) for member in configuration.members
                         if member.field and member.field.is_active and member.field.deleted_at is None]
        if templates and any(
            _normalized_name(name) == requested_name
            and _key(template_type).startswith(_key(field_type))
            for name, template_type in templates
        ):
            candidates[layout_code] = configured_by_code.get(layout_code)
    return candidates


def build_preview(db, season_id, raw_rows):
    season = db.get(Season, season_id)
    weeks = {w.week_number: w for w in db.query(Week).filter(Week.season_id == season_id).all()}
    sites = {_key(x.name): x for x in db.query(HostLocation).filter(HostLocation.is_active.is_(True)).all()}
    all_divisions = db.query(Division).all()
    divisions = {_key(f'{x.division_group or ""} {x.name}'.strip()): x for x in all_divisions}
    divisions.update({_key(x.name): x for x in all_divisions})
    alias_groups = {}
    for candidate in all_divisions:
        for label in (candidate.name, f'{candidate.division_group or ""} {candidate.name}'.strip()):
            alias_groups.setdefault(_division_alias(label), []).append(candidate)
    divisions.update({key: values[0] for key, values in alias_groups.items()
                      if len({value.id for value in values}) == 1})
    teams = season_roster(db, season_id)
    status_model = __import__('app.models', fromlist=['GameStatus']).GameStatus
    scheduled_status = db.query(status_model).filter_by(code='SCHEDULED').first()
    results, staged, game_keys, field_keys, team_keys = [], [], set(), set(), set()
    timeslot_rows = {}

    for number, raw in enumerate(raw_rows, 2):
        errors = []
        warnings = []
        source_week = _text(raw.get('week'))
        week_number = _week_number(raw.get('week'))
        week = weeks.get(week_number)
        game_date, kickoff = _date(raw.get('date')), _time(raw.get('kickoff'))
        if not week:
            errors.append(f'Week "{source_week}" could not be matched to a configured season week.')
        if not game_date:
            errors.append(f'Date "{_text(raw.get("date"))}" is invalid.')
        elif week and not (week.start_date <= game_date <= week.end_date):
            configured = week.primary_game_date or week.start_date
            errors.append(f'Week {week.week_number} is configured for {configured.strftime("%B %-d, %Y")}, but the imported date is {game_date.strftime("%B %-d, %Y")}.')
        if not kickoff:
            errors.append(f'Kickoff "{_text(raw.get("kickoff"))}" is invalid.')

        site = sites.get(_key(raw.get('site')))
        if not site:
            errors.append(f'Site "{_text(raw.get("site"))}" could not be found.')
        field = field_instance = physical_area = None
        area_name, slot_name = _split_area_slot(raw)
        area_configuration_candidates = []
        configuration_candidates = []
        uses_physical_areas = False
        if site:
            areas = db.query(PhysicalFieldArea).filter_by(
                host_location_id=site.id, is_active=True).all()
            uses_physical_areas = bool(areas)
            if uses_physical_areas and area_name:
                area_matches = [item for item in areas
                                if _normalized_name(item.name) == _normalized_name(area_name)]
                physical_area = area_matches[0] if len(area_matches) == 1 else None
                if not physical_area:
                    errors.append(f'Physical Area "{area_name}" was not found at {site.name}.')
                else:
                    area_configuration_candidates = _area_slot_candidates(
                        db, physical_area, slot_name, _normalized_field_type(raw.get('fieldtype')))
                    if not area_configuration_candidates:
                        errors.append(f'Generated slot "{slot_name}" is not supported by {physical_area.name}.')
            elif uses_physical_areas and _slot_signature(slot_name, _normalized_field_type(raw.get('fieldtype'))):
                supporting = [(area, _area_slot_candidates(
                    db, area, slot_name, _normalized_field_type(raw.get('fieldtype')))) for area in areas]
                supporting = [(area, choices) for area, choices in supporting if choices]
                if len(supporting) == 1:
                    physical_area, area_configuration_candidates = supporting[0]
                    area_name = physical_area.name
                elif len(supporting) > 1:
                    errors.append(f'Generated slot "{slot_name}" is ambiguous at {site.name}; add a Physical Area column or use "Physical Area / Slot" in Field.')
            if not uses_physical_areas:
                field = resolve_legacy_import_field(
                    db, site, raw.get('physicalarea'), raw.get('field'))
                # In the flat architecture the Physical Area cell is another
                # possible field label, not a separate hierarchy level.
                area_name = ''
                if field:
                    slot_name = field.name
            elif not physical_area:
                field = resolve_active_field(db, site, raw.get('field'))
            instances = db.query(FieldInstance).filter(FieldInstance.host_location_id == site.id, FieldInstance.is_active.is_(True)).all()
            if not physical_area:
                field_instance = next((x for x in instances if _normalized_name(x.field_name) == _normalized_name(raw.get('field')) and (not game_date or x.instance_date == game_date)), None)
            # Generated/manual scheduling persists the dated field instance as
            # the playable assignment.  Its display name can include layout
            # terms, so prefer its canonical availability -> Field relationship
            # over trying to match that generated label to spreadsheet text.
            if field and not field_instance:
                field_instance = next((
                    x for x in instances
                    if (not game_date or x.instance_date == game_date)
                    and getattr(x.hosting_availability, 'field_id', None) == field.id
                ), None)
            # A dynamic position may exist only in a supported site layout.
            configuration_candidates = _configuration_candidates(
                db, site, field.name if field else raw.get('field'), raw.get('fieldtype'))
            if not physical_area and not field and not field_instance and not configuration_candidates and not errors:
                if not uses_physical_areas and _text(raw.get('physicalarea')):
                    missing = _text(raw.get('physicalarea')) or _text(raw.get('field'))
                    errors.append(f'Field "{missing}" was not found at {site.name}.')
                else:
                    errors.append(f'Field "{_text(raw.get("field"))}" could not be found at site "{_text(raw.get("site"))}".')

        field_type = _normalized_field_type(raw.get('fieldtype'))
        configured_field_type = None
        if not field_type:
            errors.append('Field Type must be Small, Medium, or Large.')
        else:
            # ``Field.layout_type`` is the configured type of the canonical
            # field resolved above.  A dated instance can reflect a generated
            # layout, so it is only authoritative when no canonical Field was
            # resolved.  Site configuration names (for example FOUR_SMALL)
            # are deliberately not individual field types.
            resolved_type = (field.layout_type if field else None) or (
                field_instance.field_type if field_instance else None)
            canonical_field_type = _normalized_field_type(resolved_type)
            configured_field_type = canonical_field_type.title() if canonical_field_type else None
            type_mismatch = bool(canonical_field_type and canonical_field_type != field_type)
            # A field can be deliberately reconfigured for one timeslot.  Keep
            # this advisory separate from errors so it never prevents staging.
            if type_mismatch:
                if canonical_field_type:
                    warnings.append(
                        f'Field "{_text(raw.get("field"))}" is normally configured as '
                        f'"{canonical_field_type.title()}", but the import requests '
                        f'"{field_type.title()}". Verify the field will be '
                        'reconfigured for this timeslot.'
                    )

        division = divisions.get(_key(raw.get('division'))) or divisions.get(_division_alias(raw.get('division')))
        if not division:
            errors.append(f'Division "{_text(raw.get("division"))}" could not be found.')

        home_resolution = resolve_roster_team(teams, division, raw.get('hometeam'))
        away_resolution = resolve_roster_team(teams, division, raw.get('awayteam'))
        home, away = home_resolution.team, away_resolution.team
        season_name = season.name if season else 'selected'
        def resolution_error(side, value, resolution):
            imported = _text(value)
            if resolution.candidate_count > 1:
                return (f'Multiple active teams match "{imported}". '
                        'Import cannot continue until the team data is unambiguous.')
            parsed_community = resolution.community or 'Not recognized'
            parsed_team = resolution.team_name or 'Not recognized'
            return (f'Unable to resolve {side.lower()} team "{imported}". Parsed: '
                    f'Community: {parsed_community}; Division: {_text(raw.get("division"))}; '
                    f'Team: {parsed_team}; Season: {season_name}. '
                    'No matching active team was found.')
        if not home:
            errors.append(resolution_error('Home', raw.get('hometeam'), home_resolution))
        if not away:
            errors.append(resolution_error('Away', raw.get('awayteam'), away_resolution))
        if home and away and home.id == away.id:
            errors.append('Home Team and Away Team cannot be the same.')

        if week and game_date and kickoff and site and (physical_area or field or field_instance or configuration_candidates) and home and away:
            identity = ((physical_area.id, _normalized_name(slot_name)) if physical_area else
                        field.id if field else (field_instance.id if field_instance else _normalized_name(raw.get('field'))))
            game_key = (week.id, game_date, kickoff, site.id, identity, home.id, away.id)
            field_key = (game_date, kickoff, site.id, identity)
            simultaneous = {(game_date, kickoff, home.id), (game_date, kickoff, away.id)}
            if game_key in game_keys:
                errors.append('Duplicate imported game.')
            if field_key in field_keys:
                errors.append('Field conflict at this date and kickoff.')
            if team_keys.intersection(simultaneous):
                errors.append('Team conflict at this date and kickoff.')
            game_keys.add(game_key); field_keys.add(field_key); team_keys.update(simultaneous)

        row = {'row': number, 'week': f'Week {week_number}' if week_number else source_week,
               'date': game_date.isoformat() if game_date else _text(raw.get('date')),
               'kickoff': kickoff.strftime('%H:%M') if kickoff else _text(raw.get('kickoff')),
               'site': _text(raw.get('site')), 'physical_area': area_name or None,
               'field': slot_name if physical_area else (field.name if field else _text(raw.get('field'))),
               'field_architecture': 'physical_area' if uses_physical_areas else 'legacy_field',
               'configuration': _text(raw.get('layout')) or None,
               'configured_field_type': configured_field_type,
               'imported_field_type': _text(raw.get('fieldtype')).title(),
               'division': _text(raw.get('division')), 'home_team': _text(raw.get('hometeam')),
               'away_team': _text(raw.get('awayteam')), 'notes': _text(raw.get('notes')) or None,
               'status': 'ERROR' if errors else ('WARNING' if warnings else 'VALID'),
               'message': ' '.join(errors + warnings) or 'Ready to import.'}
        results.append(row)
        if not errors:
            resolved_slot = None
            if physical_area and field_instance and game_date and kickoff:
                resolved_slot = db.query(GameSlot).filter(
                    GameSlot.field_instance_id == field_instance.id,
                    GameSlot.slot_date == game_date,
                    GameSlot.start_time == kickoff,
                ).first()
            staged_row = {**row, 'week_number': week_number, 'week_id': str(week.id), 'site_id': str(site.id),
                           # Keep the source label for audit/preview, and keep the
                           # canonical relationship separately for confirmation.
                           # Confirmation must never need to resolve a name after
                           # the schedule being replaced has been deleted.
                           'imported_field_name': _text(raw.get('field')),
                           'field_architecture': 'physical_area' if uses_physical_areas else 'legacy_field',
                           'physical_area_id': str(physical_area.id) if physical_area else None,
                           'physical_area': physical_area.name if physical_area else None,
                           'resolved_field_id': str(field.id) if field else None,
                           'field_id': str(field.id) if field else None,
                           'field_layout_type_override': field_type if type_mismatch else None,
                           'field_instance_id': str(field_instance.id) if physical_area and field_instance else None,
                           'game_slot_id': str(resolved_slot.id) if resolved_slot else None,
                           'home_team_id': str(home.id), 'away_team_id': str(away.id),
                           'game_status_id': str(scheduled_status.id) if scheduled_status else None}
            staged.append(staged_row)
            if physical_area and game_date and kickoff:
                key = (site.id, physical_area.id, game_date, kickoff)
                timeslot_rows.setdefault(key, []).append((
                    row, staged_row, {str(option.id): option for option in area_configuration_candidates}
                ))
            if configuration_candidates and game_date and kickoff:
                key = (site.id, game_date, kickoff)
                timeslot_rows.setdefault(key, []).append((
                    row, staged_row, configuration_candidates
                ))

    # A layout is a property of the complete site/date/kickoff wave, not of an
    # individual field record. Intersecting candidates rejects combinations
    # whose physical footprints overlap (for example Hiller Field 1 / Large and
    # Field 2 / Medium), even though each assignment is possible by itself.
    invalid_staged_ids = set()
    for group_key, grouped in timeslot_rows.items():
        kickoff = group_key[-1]
        is_area_group = len(group_key) == 4
        site = db.get(HostLocation, grouped[0][1]['site_id'])
        if is_area_group:
            sample_staged = grouped[0][1]
            area = db.get(PhysicalFieldArea, sample_staged['physical_area_id'])
            game_date = _date(grouped[0][0]['date'])
            availability = _area_hosting_availability(
                db, season_id, site, area, game_date, kickoff)
            if not availability:
                message = (
                    f'{area.name} is not marked available for hosting on '
                    f'{game_date.strftime("%m/%d/%Y")} at {kickoff.strftime("%I:%M %p")}.'
                )
                for row, staged_row, _candidates in grouped:
                    row['status'] = 'ERROR'; row['message'] = message
                    invalid_staged_ids.add(id(staged_row))
                continue
        common_ids = set(grouped[0][2])
        for _row, _staged_row, candidates in grouped[1:]:
            common_ids.intersection_update(candidates)
        if not common_ids:
            site_name = grouped[0][0]['site']
            area_label = grouped[0][0].get('physical_area')
            slots = ' + '.join(row['field'] for row, _staged, _choices in grouped)
            message = (f'The imported slots {slots} do not match any supported layout for '
                       f'{area_label or site_name} at {kickoff.strftime("%-I:%M %p")}.')
            for row, staged_row, _candidates in grouped:
                row['status'] = 'ERROR'; row['message'] = message
                invalid_staged_ids.add(id(staged_row))
        else:
            # Prefer an exact slot-count layout; a merely larger configuration
            # is not valid inference for an incomplete imported wave.
            imported_counts = {size: sum(1 for row, *_ in grouped
                                          if _normalized_field_type(row['imported_field_type']) == size)
                               for size in ('SMALL', 'MEDIUM', 'LARGE')}
            exact_ids = common_ids
            if is_area_group:
                def inference_rank(code):
                    option = grouped[0][2][code]
                    capacity = {size: int(getattr(option, f'{size.lower()}_field_count', 0) or 0)
                                for size in imported_counts}
                    return (sum(capacity[size] for size, count in imported_counts.items() if not count),
                            sum(capacity.values()) - sum(imported_counts.values()))
                best_rank = min(map(inference_rank, common_ids))
                exact_ids = {code for code in common_ids if inference_rank(code) == best_rank}
            explicit_labels = {_normalized_name(row.get('configuration'))
                               for row, *_ in grouped if _text(row.get('configuration'))}
            if len(explicit_labels) > 1:
                message = (f'Conflicting configurations were specified for '
                           f'{grouped[0][0].get("physical_area") or grouped[0][0]["site"]} '
                           f'at {kickoff.strftime("%-I:%M %p")}. All games using the same '
                           'physical area during a timeslot must use the same field layout.')
                for row, staged_row, _ in grouped:
                    row['status'] = 'ERROR'; row['message'] = message
                    invalid_staged_ids.add(id(staged_row))
                continue
            explicit = _text(grouped[0][0].get('configuration'))
            if explicit:
                exact_ids = {code for code in common_ids if
                             _normalized_name(getattr(grouped[0][2][code], 'name', None)
                                              or grouped[0][2][code].configuration_name) == _normalized_name(explicit)}
            elif not is_area_group:
                exact_ids = {sorted(common_ids)[0]}
            if len(exact_ids) != 1:
                slots = ' + '.join(row['field'] for row, *_ in grouped)
                message = (f'The imported slots {slots} do not match any supported layout for '
                           f'{grouped[0][0].get("physical_area") or grouped[0][0]["site"]} '
                           f'at {kickoff.strftime("%-I:%M %p")}.' if not exact_ids else
                           f'The imported slots are ambiguous; specify Layout for {grouped[0][0].get("physical_area")} at {kickoff.strftime("%-I:%M %p")}.')
                for row, staged_row, _ in grouped:
                    row['status'] = 'ERROR'; row['message'] = message
                    invalid_staged_ids.add(id(staged_row))
                continue
            selected_code = next(iter(exact_ids))
            # An already materialized physical-area availability is the
            # authoritative layout for this wave.  Prefer it when its actual
            # generated slots cover this import, even if inference selected a
            # smaller layout.  This both preserves capacity and stages the
            # stable instance/slot IDs belonging to the saved layout.
            if is_area_group and availability.field_configuration_option_id:
                existing_code = str(availability.field_configuration_option_id)
                if existing_code in common_ids:
                    required_slots = []
                    for row, _staged_row, _candidates in grouped:
                        instance = _find_area_instance(
                            db, site, area, availability.field_configuration_option,
                            row['field'], game_date, kickoff)
                        slot = (db.query(GameSlot).filter_by(
                            field_instance_id=instance.id, slot_date=game_date,
                            start_time=kickoff).first() if instance else None)
                        if slot:
                            required_slots.append(slot.id)
                    available_slots = {
                        slot.id for slot in db.query(GameSlot).join(
                            FieldInstance, GameSlot.field_instance_id == FieldInstance.id
                        ).filter(
                            FieldInstance.hosting_availability_id == availability.id,
                            GameSlot.slot_date == game_date,
                            GameSlot.start_time == kickoff,
                        ).all()
                    }
                    if configuration_supports_required_slots(
                            available_slots, required_slots) and len(required_slots) == len(grouped):
                        selected_code = existing_code
            selected = grouped[0][2][selected_code]
            layout = getattr(selected, 'name', None) or selected.configuration_name.replace('_', ' ').title()
            if not is_area_group and (site.surface_type or '').upper() != 'TURF_STADIUM':
                integrity_error = _layout_integrity_error(selected, site)
                if integrity_error:
                    message = f'Configuration "{layout}" for {site.name} is invalid. {integrity_error}'
                    for row, staged_row, _ in grouped:
                        row['status'] = 'ERROR'; row['message'] = message
                        invalid_staged_ids.add(id(staged_row))
                    continue
            resolved_member_ids = ([str(member.field_id) for member in selected.members]
                                   if not is_area_group else [])
            for row, staged_row, _candidates in grouped:
                # This is the canonical scheduling-block identity used by
                # confirmation.  Persist it in staging so commit never has to
                # infer a layout again from display labels.
                staged_row['configuration_group_key'] = '|'.join(str(value) for value in (
                    season_id, staged_row['date'], staged_row['kickoff'],
                    staged_row['site_id'],
                    staged_row.get('physical_area_id') if is_area_group else '',
                ))
                staged_row['resolved_configuration_architecture'] = (
                    'physical_area' if is_area_group else 'legacy_layout')
                if row.get('physical_area'):
                    staged_row['field_configuration_option_id'] = str(selected.id)
                    staged_row['area_configuration_name'] = layout
                    row['configuration'] = layout
                    area = db.get(PhysicalFieldArea, staged_row['physical_area_id'])
                    group_site = db.get(HostLocation, staged_row['site_id'])
                    instance = _find_area_instance(db, group_site, area, selected, row['field'],
                                                   _date(row['date']), _time(row['kickoff']))
                    if instance:
                        staged_row['field_instance_id'] = str(instance.id)
                        # Physical-area slots are generated resources rather
                        # than permanent ``Field`` rows.  Persist the complete
                        # generated assignment selected by the grouped layout:
                        # its dated instance and its exact kickoff slot.
                        slot = db.query(GameSlot).filter_by(
                            field_instance_id=instance.id,
                            slot_date=_date(row['date']),
                            start_time=_time(row['kickoff']),
                        ).first()
                        if slot:
                            staged_row['game_slot_id'] = str(slot.id)
                            availability_field_id = getattr(
                                instance.hosting_availability, 'field_id', None)
                            if availability_field_id:
                                staged_row['resolved_field_id'] = str(availability_field_id)
                                staged_row['field_id'] = str(availability_field_id)
                    # Runtime instances/slots are derived resources.  Their
                    # absence during preview is valid; confirmation
                    # materializes them from this staged capability assignment.
                else:
                    staged_row['configuration_id'] = str(selected.id) if selected else None
                    staged_row['configuration_name'] = selected_code
                    staged_row['layout_field_ids'] = resolved_member_ids
                layout_message = f'{row.get("physical_area") or row["site"]} will use its {layout} configuration for this timeslot.'
                row['message'] = (f'{row["message"]} {layout_message}' if row['status'] == 'WARNING'
                                  else f'Ready to import. {layout_message}')
            logger.info(
                'schedule_import_timeslot_resolved host_location_id=%s physical_area_id=%s '
                'physical_area_name=%s date=%s kickoff=%s layout_id=%s layout_name=%s '
                'resolved_field_ids=%s source_rows=%s',
                grouped[0][1]['site_id'], grouped[0][1].get('physical_area_id'),
                grouped[0][0].get('physical_area'), grouped[0][0]['date'],
                grouped[0][0]['kickoff'], selected.id, layout, resolved_member_ids,
                [row['row'] for row, *_ in grouped],
            )

    # A persisted legacy layout is authoritative.  Compare its canonical field
    # membership with all assignments in the imported wave; layout labels are
    # intentionally irrelevant and unused saved capacity is compatible.
    legacy_groups = {}
    for staged_row in staged:
        if (id(staged_row) not in invalid_staged_ids
                and staged_row.get('resolved_configuration_architecture') == 'legacy_layout'):
            key = (staged_row['site_id'], staged_row['date'], staged_row['kickoff'])
            legacy_groups.setdefault(key, []).append(staged_row)
    for (site_id, configuration_date, kickoff_value), grouped_rows in legacy_groups.items():
        existing = db.query(TimeslotFieldConfiguration).filter_by(
            host_location_id=site_id, configuration_date=_date(configuration_date),
            kickoff_time=_time(kickoff_value)).first()
        if not existing:
            continue
        existing_integrity_error = _layout_integrity_error(
            existing.configuration, db.get(HostLocation, site_id))
        if existing_integrity_error:
            message = (f'Existing configuration "{layout_label(existing.configuration)}" is invalid. '
                       f'{existing_integrity_error}')
            for staged_row in grouped_rows:
                matching_result = next(item for item in results if item['row'] == staged_row['row'])
                matching_result['status'] = 'ERROR'; matching_result['message'] = message
                invalid_staged_ids.add(id(staged_row))
            continue
        available_ids = {str(member.field_id) for member in existing.configuration.members
                         if member.field and member.field.is_active and member.field.deleted_at is None}
        required_ids = {row.get('resolved_field_id') or row.get('field_id') for row in grouped_rows}
        required_ids.discard(None)
        if configuration_supports_required_slots(available_ids, required_ids):
            current_label = layout_label(existing.configuration)
            for staged_row in grouped_rows:
                staged_row['configuration_id'] = str(existing.configuration_id)
                staged_row['configuration_name'] = existing.configuration.configuration_name
                staged_row['layout_field_ids'] = sorted(available_ids)
                matching_result = next(item for item in results if item['row'] == staged_row['row'])
                matching_result['configuration'] = current_label
                matching_result['message'] = (
                    f'Ready to import. Existing configuration "{current_label}" will be reused.')
            continue
        missing_ids = required_ids - available_ids
        missing_fields = [db.get(Field, field_id) for field_id in missing_ids]
        missing_label = ', '.join(f'"{field.name}"' for field in missing_fields if field)
        message = (f'Existing configuration "{layout_label(existing.configuration)}" does not '
                   f'provide required field {missing_label or "assignment"}.')
        for staged_row in grouped_rows:
            matching_result = next(item for item in results if item['row'] == staged_row['row'])
            matching_result['status'] = 'ERROR'; matching_result['message'] = message
            invalid_staged_ids.add(id(staged_row))
    staged = [row for row in staged if id(row) not in invalid_staged_ids]

    affected = sorted({x['week_number'] for x in staged})
    existing = db.query(Game).join(Week, Game.week_id == Week.id).filter(
        Game.season_id == season_id, Week.week_number.in_(affected)).count() if affected else 0
    regular = {w.week_number for w in weeks.values() if w.date_type == 'REGULAR_SEASON'}
    dates = [_date(x['date']) for x in staged]
    warning_rows = [row for row in results if row['status'] == 'WARNING']
    diagnostics = [{
        'category': 'Scheduling Integrity',
        'check': 'Field configuration mismatch',
        'status': 'WARNING',
        'advisory': True,
        'blocking': False,
        'detail': {
            'site': row['site'], 'kickoff': row['kickoff'], 'field': row['field'],
            'configured_default_type': row['configured_field_type'],
            'imported_type': row['imported_field_type'],
            'message': row['message'],
        },
    } for row in warning_rows]
    return {'rows_uploaded': len(raw_rows), 'valid_games': len(staged),
            'importable_games': len(staged), 'invalid_rows': len(results) - len(staged),
            'weeks': affected, 'earliest_date': min(dates).isoformat() if dates else None,
            'latest_date': max(dates).isoformat() if dates else None, 'sites': sorted({x['site'] for x in staged}),
            'divisions': sorted({x['division'] for x in staged}), 'games_to_add': len(staged),
            'existing_games_to_replace': existing, 'warning_count': len(warning_rows),
            'blocking_errors': len(results) - len(staged),
            'is_full_regular_season': bool(regular) and set(affected) == regular,
            'diagnostics': diagnostics, 'rows': results}, staged
